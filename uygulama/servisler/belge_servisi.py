#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Belge Oluşturma Servisi — Excel şablon birleştirme motoru.
"""

import os
import re
import hashlib
import shutil
from copy import copy
from typing import Tuple, Optional

from uygulama.altyapi.belge_repo import BelgeRepository
from uygulama.ortak.yardimcilar import logger_olustur, kullanici_veri_dizini
from uygulama.ortak.app_state import app_state

logger = logger_olustur("belge_srv")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError as import_err:
    HAS_OPENPYXL = False
    logger.warning(
        f"UYARI: Excel kütüphanesi (openpyxl) yüklü değil.\n"
        f"Hata detayı: {import_err}\n"
        f"Belge oluşturma özellikleri kullanılamayacaktır.\n"
        f"Çözüm: pip install openpyxl==3.1"
    )


def _col_parse(aralik: str) -> Tuple[int, int]:
    parts = aralik.strip().upper().split(":")
    if len(parts) == 2:
        return column_index_from_string(parts[0]), column_index_from_string(parts[1])
    return 1, 9


def _phash(pid: str) -> str:
    return hashlib.md5(pid.encode()).hexdigest()[:8].upper()


class BelgeServisi:
    def __init__(self, belge_repo: BelgeRepository,
                 teklif_srv=None, placeholder_srv=None,
                 proje_srv=None, em_repo=None):
        self.repo = belge_repo
        self.teklif_srv = teklif_srv
        self.ph_srv = placeholder_srv
        self.proje_srv = proje_srv
        self.em_repo = em_repo

    # ═══════════════════════════════════════
    # ŞABLON DOSYA YÖNETİMİ
    # ═══════════════════════════════════════

    def sablon_yukle(self, kaynak_yol: str, ad: str,
                      sheet_adi: str = "Sheet1") -> Tuple[bool, str, Optional[str]]:
        if not os.path.exists(kaynak_yol):
            return False, "Dosya bulunamadı.", None
        sablon_dir = os.path.join(kullanici_veri_dizini(), "sablonlar")
        os.makedirs(sablon_dir, exist_ok=True)
        dosya_adi = os.path.basename(kaynak_yol)
        hedef = os.path.join(sablon_dir, dosya_adi)
        i = 1
        while os.path.exists(hedef):
            base, ext = os.path.splitext(dosya_adi)
            hedef = os.path.join(sablon_dir, f"{base}_{i}{ext}")
            i += 1
        shutil.copy2(kaynak_yol, hedef)
        sid = self.repo.sablon_dosya_ekle(ad, hedef, sheet_adi)
        return True, f"Şablon yüklendi: {ad}", sid

    def sablon_sheetleri(self, dosya_yolu: str) -> list[str]:
        if not HAS_OPENPYXL or not os.path.exists(dosya_yolu):
            if not os.path.exists(dosya_yolu):
                logger.debug(f"Şablon dosyası yok: {dosya_yolu}")
            return []
        try:
            wb = load_workbook(dosya_yolu, read_only=True)
            s = wb.sheetnames
            wb.close()
            return s
        except PermissionError as perm_err:
            logger.error(f"Şablon dosyası okunamadı (İzin yok): {dosya_yolu}\nHata: {perm_err}")
            return []
        except Exception as e:
            logger.error(f"Şablon dosyası okunamadı: {dosya_yolu}\nHata tipi: {type(e).__name__}\nDetay: {e}")
            return []

    # ═══════════════════════════════════════
    # BELGE OLUŞTURMA MOTORU
    # ═══════════════════════════════════════

    def belge_olustur(self, teklif_id: str, belge_turu_kodu: str,
                       hedef_klasor: str) -> Tuple[bool, str, Optional[str]]:
        if not HAS_OPENPYXL:
            return False, "openpyxl kütüphanesi yüklü değil.", None

        teklif = self.teklif_srv.getir(teklif_id) if self.teklif_srv else None
        if not teklif:
            logger.error(f"Belge oluşturulamadı: Teklif ID '{teklif_id}' bulunamadı")
            return False, "Teklif bulunamadı.", None

        proje = self.proje_srv.getir(teklif["proje_id"]) if self.proje_srv else None
        if not proje:
            logger.error(f"Belge oluşturulamadı: Teklif {teklif_id} için Proje '{teklif['proje_id']}' bulunamadı")
            return False, "Proje bulunamadı.", None

        belge_turu = self.repo.belge_turu_kod_ile(belge_turu_kodu)
        if not belge_turu:
            logger.error(f"Belge oluşturulamadı: Belge türü kodu '{belge_turu_kodu}' bulunamadı")
            return False, f"Belge türü bulunamadı: {belge_turu_kodu}", None

        proje_urunler = self.proje_srv.proje_urunleri(
            teklif["proje_id"]) if self.proje_srv else []

        c1, c2 = _col_parse(belge_turu["sutun_araligi"])
        bolumler = self.repo.bolumler(belge_turu["id"])
        if not bolumler:
            logger.error(f"Belge oluşturulamadı: Belge türü '{belge_turu_kodu}' ({belge_turu['id']}) için bölüm tanımlı değil")
            return False, "Bu belge türü için bölüm tanımlı değil.", None

        wb_out = Workbook(); ws_out = wb_out.active; ws_out.title = "Belge"
        out_row = 1
        proje_bilgi = self._proje_baglam(proje)
        teklif_params = self._teklif_param_baglam(teklif_id)

        for bolum in bolumler:
            for ab in self._bolum_atamalari(bolum, proje_urunler, teklif_id):
                atama = ab["atama"]
                bglm = ab.get("baglam", {})
                tam = {
                    "proje_bilgi": proje_bilgi,
                    "teklif_param": {**teklif_params, **bglm.get("teklif_param", {})},
                    "urun_param": bglm.get("urun_param", {}),
                    "alt_kalem_param": bglm.get("alt_kalem_param", {}),
                }
                n = self._satirlari_kopyala(
                    ws_out, out_row, atama["dosya_yolu"], atama["sheet_adi"],
                    atama["satir_baslangic"], atama["satir_bitis"], c1, c2)
                if self.ph_srv and n > 0:
                    self._ph_degistir(ws_out, out_row, out_row + n - 1, c1, c2, tam)
                out_row += n

        self._sutun_genislikleri(ws_out, bolumler, c1, c2)

        # Dosya/klasör adları
        p_ad = self._p(proje, "firma")
        p_konum = self._p(proje, "konum")
        p_id = self._p(proje, "id")
        ph = _phash(p_id)
        urun_str = "_".join(self._urun_kodlari(proje_urunler)) or "URUNLER"
        safe = lambda s: re.sub(r'[<>:"/\\|?*]', '_', s)

        klasor_adi = safe(f"{p_ad} {p_konum} {urun_str} {ph}".strip())
        klasor_yolu = os.path.join(hedef_klasor, klasor_adi)
        
        try:
            os.makedirs(klasor_yolu, exist_ok=True)
        except Exception as mkdir_err:
            logger.error(f"Klasör oluşturulamadı: {klasor_yolu}\nHata: {type(mkdir_err).__name__} - {mkdir_err}")
            return False, f"Klasör oluşturulamadı: {klasor_yolu}\nHata: {mkdir_err}", None

        dosya_adi = safe(f"{p_ad} {p_konum} {urun_str} {ph}.xlsx".strip())
        dosya_yolu = os.path.join(klasor_yolu, dosya_adi)

        try:
            wb_out.save(dosya_yolu)
        except PermissionError as perm_err:
            logger.error(f"Dosya kaydetme hatası (İzin Reddedildi): {dosya_yolu}\nHata: {perm_err}")
            return False, f"Dosya kaydедilemiyor - yeterli izin yok.\nLütfen dosyanın kapalı olduğundan emin olun.", None
        except IOError as io_err:
            logger.error(f"Dosya I/O hatası: {dosya_yolu}\nHata: {type(io_err).__name__} - {io_err}")
            return False, f"Dosya yazma hatası: {io_err}", None
        except Exception as save_err:
            logger.error(f"Belge kaydetme başarısız: {dosya_yolu}\nHata tipi: {type(save_err).__name__}\nHata: {save_err}")
            return False, f"Kaydetme hatası: {type(save_err).__name__} - {save_err}", None

        state = app_state()
        olusan = state.aktif_kullanici.kullanici_adi if state.aktif_kullanici else ""
        try:
            self.repo.uretim_kaydet(
                teklif_id, belge_turu["id"], dosya_yolu, dosya_adi, klasor_yolu, olusan)
        except Exception as db_err:
            logger.warning(f"Belge oluşturuldu ama veritabanına kaydedilemedi:\nDosya: {dosya_yolu}\nHata: {type(db_err).__name__} - {db_err}")
        
        logger.info(f"Belge oluşturuldu: {dosya_yolu}")
        return True, f"Belge oluşturuldu: {dosya_adi}", dosya_yolu

    # ═══════════════════════════════════════
    # BÖLÜM ATAMALARI
    # ═══════════════════════════════════════

    def _bolum_atamalari(self, bolum, proje_urunler, teklif_id):
        raw = self.repo.atamalar(bolum["id"])
        sonuc = []
        if bolum["tur"] == "sabit":
            for a in raw:
                sonuc.append({"atama": a, "baglam": {}})

        elif bolum["tur"] == "urun_bazli":
            for pu in proje_urunler:
                uid = pu["urun_id"]
                for a in raw:
                    if a.get("urun_id") == uid:
                        sonuc.append({"atama": a,
                                      "baglam": self._urun_baglam(uid, teklif_id)})

        elif bolum["tur"] == "alt_kalem_bazli":
            kalemler = self.teklif_srv.zenginlestirilmis_kalemler(teklif_id) if self.teklif_srv else []
            for pu in proje_urunler:
                uid = pu["urun_id"]
                for k in kalemler:
                    if k["urun_id"] != uid or k["tip"] != "alt_kalem":
                        continue
                    if k.get("dahil_durumu", "DAHIL") == "HARIC":
                        continue
                    akid = k.get("alt_kalem_id")
                    for a in raw:
                        if a.get("urun_id") == uid and a.get("alt_kalem_id") == akid:
                            sonuc.append({"atama": a,
                                          "baglam": self._ak_baglam(uid, akid, k["id"], teklif_id)})

        elif bolum["tur"] == "urun_alt_kalem":
            # Her ürün için: önce ürün başlığı (alt_kalem_id=NULL), sonra alt kalemler
            kalemler = self.teklif_srv.zenginlestirilmis_kalemler(teklif_id) if self.teklif_srv else []
            for pu in proje_urunler:
                uid = pu["urun_id"]
                # Ürün başlık ataması (alt_kalem_id NULL)
                for a in raw:
                    if a.get("urun_id") == uid and not a.get("alt_kalem_id"):
                        sonuc.append({"atama": a,
                                      "baglam": self._urun_baglam(uid, teklif_id)})
                # Alt kalem atamaları
                for k in kalemler:
                    if k["urun_id"] != uid or k["tip"] != "alt_kalem":
                        continue
                    if k.get("dahil_durumu", "DAHIL") == "HARIC":
                        continue
                    akid = k.get("alt_kalem_id")
                    for a in raw:
                        if a.get("urun_id") == uid and a.get("alt_kalem_id") == akid:
                            sonuc.append({"atama": a,
                                          "baglam": self._ak_baglam(uid, akid, k["id"], teklif_id)})

        return sonuc

    # ═══════════════════════════════════════
    # BAĞLAM
    # ═══════════════════════════════════════

    def _proje_baglam(self, proje) -> dict:
        g = lambda k: getattr(proje, k, "") if not isinstance(proje, dict) else proje.get(k, "")
        return {f"PROJE_{k.upper()}": g(k) for k in ["firma", "konum", "tesis", "urun_seti"]}

    def _teklif_param_baglam(self, teklif_id: str) -> dict:
        """Tüm teklif kalemlerinin özel parametrelerini toplar."""
        if not self.teklif_srv: return {}
        p = {}
        for k in self.teklif_srv.zenginlestirilmis_kalemler(teklif_id):
            for v in self.teklif_srv.parametre_degerleri(k["id"]):
                n = v["parametre_adi"]
                # Özel/spesifik parametreleri al (__ prefix veya tamamen UPPER)
                if n.startswith("__") or n == n.upper():
                    p[n] = v["deger"]
        return p

    def _urun_baglam(self, urun_id, teklif_id):
        bg = {"urun_param": {}, "teklif_param": {}}
        if self.em_repo:
            ver = self.em_repo.aktif_urun_versiyon(urun_id)
            if ver:
                for p in self.em_repo.urun_parametreleri(ver["id"]):
                    bg["urun_param"][p["ad"]] = p["varsayilan_deger"]
        if self.teklif_srv:
            for k in self.teklif_srv.zenginlestirilmis_kalemler(teklif_id):
                if k["urun_id"] == urun_id and k["tip"] == "urun":
                    for v in self.teklif_srv.parametre_degerleri(k["id"]):
                        if not v["parametre_adi"].startswith("_"):
                            bg["urun_param"][v["parametre_adi"]] = v["deger"]
        return bg

    def _ak_baglam(self, urun_id, ak_id, kalem_id, teklif_id):
        bg = {"alt_kalem_param": {}, "urun_param": {}, "teklif_param": {}}
        if self.teklif_srv:
            for v in self.teklif_srv.parametre_degerleri(kalem_id):
                n = v["parametre_adi"]
                if n.startswith("__") or n == n.upper():
                    bg["teklif_param"][n] = v["deger"]
                elif not n.startswith("_"):
                    bg["alt_kalem_param"][n] = v["deger"]
        ub = self._urun_baglam(urun_id, teklif_id)
        bg["urun_param"] = ub.get("urun_param", {})
        return bg

    # ═══════════════════════════════════════
    # EXCEL İŞLEMLERİ
    # ═══════════════════════════════════════

    def _satirlari_kopyala(self, ws_out, out_row, dosya, sheet,
                             r1, r2, c1, c2) -> int:
        if not os.path.exists(dosya):
            logger.warning(f"Şablon dosyası bulunamadı: {dosya}")
            return 0
        try:
            wb = load_workbook(dosya)
        except PermissionError as perm_err:
            logger.error(f"Şablon dosyası okunamadı (İzin yok): {dosya}\nHata: {perm_err}")
            return 0
        except Exception as e:
            logger.error(f"Şablon dosyası açılamadı: {dosya}\nHata tipi: {type(e).__name__}\nDetay: {e}")
            return 0
        
        if sheet not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet}' bulunamadı. Dosya: {dosya}\nMevcut sheets: {wb.sheetnames}")
            wb.close()
            return 0
        
        ws = wb[sheet]
        n = 0
        for sr in range(r1, r2 + 1):
            dr = out_row + n
            if ws.row_dimensions[sr].height:
                ws_out.row_dimensions[dr].height = ws.row_dimensions[sr].height
            for col in range(c1, c2 + 1):
                sc = ws.cell(row=sr, column=col)
                dc = ws_out.cell(row=dr, column=col)
                dc.value = sc.value
                if sc.has_style:
                    dc.font = copy(sc.font)
                    dc.border = copy(sc.border)
                    dc.fill = copy(sc.fill)
                    dc.number_format = sc.number_format
                    dc.protection = copy(sc.protection)
                    dc.alignment = copy(sc.alignment)
            n += 1
        
        # Merged cells
        for mg in ws.merged_cells.ranges:
            if mg.min_row >= r1 and mg.max_row <= r2 and mg.min_col >= c1 and mg.max_col <= c2:
                off = out_row - r1
                rng = f"{get_column_letter(mg.min_col)}{mg.min_row+off}:{get_column_letter(mg.max_col)}{mg.max_row+off}"
                try:
                    ws_out.merge_cells(rng)
                except Exception as merge_err:
                    logger.debug(f"Merged cell oluşturulamadı: {rng}\nHata: {type(merge_err).__name__} - {merge_err}")
        
        wb.close()
        return n

    def _ph_degistir(self, ws, r1, r2, c1, c2, baglam):
        pat = re.compile(r'\{/[A-ZÇĞİÖŞÜa-zçğıöşü0-9_ :]+/\}')
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cl = ws.cell(row=r, column=c)
                if cl.value and isinstance(cl.value, str) and pat.search(cl.value):
                    cl.value = self.ph_srv.toplu_cozumle(cl.value, baglam)

    def _sutun_genislikleri(self, ws_out, bolumler, c1, c2):
        for b in bolumler:
            ats = self.repo.atamalar(b["id"])
            if not ats: continue
            a = ats[0]
            if not os.path.exists(a["dosya_yolu"]): continue
            try:
                wb = load_workbook(a["dosya_yolu"])
                ws = wb[a["sheet_adi"]] if a["sheet_adi"] in wb.sheetnames else wb.active
                for col in range(c1, c2 + 1):
                    lt = get_column_letter(col)
                    if ws.column_dimensions[lt].width:
                        ws_out.column_dimensions[lt].width = ws.column_dimensions[lt].width
                wb.close(); return
            except Exception:
                continue

    # ═══════════════════════════════════════
    # YARDIMCI
    # ═══════════════════════════════════════

    def _p(self, proje, key):
        if isinstance(proje, dict): return proje.get(key, "")
        return getattr(proje, key, "")

    def _urun_kodlari(self, proje_urunler):
        kodlar = []
        if self.em_repo:
            for pu in proje_urunler:
                u = self.em_repo.db.getir_tek(
                    "SELECT kod FROM urunler WHERE id=?", (pu["urun_id"],))
                if u: kodlar.append(u["kod"])
        return kodlar

    # ═══════════════════════════════════════
    # ESKİ UYUMLULUK (proje_detay_sayfa)
    # ═══════════════════════════════════════

    def proje_belge_istatistikleri(self, proje_id: str) -> dict:
        """Geriye uyumlu — eski doküman istatistikleri stub."""
        return {"toplam_belge": 0, "onaylanan": 0, "bekleyen": 0, "toplam_maliyet": 0}
